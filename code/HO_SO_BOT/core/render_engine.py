from pathlib import Path
import datetime as dt
from typing import Dict, List, Any, Tuple
import re
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from copy import copy

from .config import TEMPLATES_DIR, DRAFT_DIR, FINAL_DIR
from .utils import format_date_vi, money_fmt, money_to_words_vi, proper_case_vn
from .naming import so_hd_full

SUFFIX_CHILD_LIST = ". Bao gồm các vật tư chi tiết:"

class RenderError(Exception):
    pass

# ---- Header & dates ----
def enrich_dates(h: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(h or {})
    for k, v in list(out.items()):
        if k.startswith("NGAY_") and v:
            d = dt.datetime.fromisoformat(v).date()
            out[k + "_FULL"] = format_date_vi(d)
            out[k + "_ISO"] = d.isoformat()
            out[k] = out[k + "_FULL"]
    return out

def build_money(header: Dict[str, Any], services: List[Dict]) -> Dict[str, Any]:
    subtotal = sum((s.get("SO_LUONG",0) or 0) * (s.get("DON_GIA",0) or 0) for s in (services or []))
    vat_percent = int(header.get("VAT", 8) or 8)
    vat = subtotal * vat_percent / 100.0
    total = subtotal + vat
    return {
        "tien_truoc_vat": money_fmt(subtotal),
        "tien_vat": money_fmt(vat),
        "tong_cong": money_fmt(total),
        "tong_cong_chu": money_to_words_vi(total),
        "VAT_PERCENT": vat_percent,
        "TONG_TRUOC_THUE": money_fmt(subtotal),
        "VAT_TIEN": money_fmt(vat),
        "TONG_CONG": money_fmt(total),
        "TONG_CONG_CHU": money_to_words_vi(total),
    }

def _fmt_money(n): 
    return money_fmt(n) if n not in (None, "") else ""

# ---- Rows (đã chốt) ----
def build_rows(services: List[Dict], materials: List[Dict]):
    services = services or []
    materials = materials or []
    # HĐ (chỉ dịch vụ, có giá)
    contract_rows = []
    for i, s in enumerate(services, start=1):
        sl = s.get("SO_LUONG", 0) or 0
        dg = s.get("DON_GIA", 0) or 0
        tt = sl * dg
        contract_rows.append({
            "stt": i,
            "ten": s.get("TEN_DV",""),
            "dvt": s.get("DVT",""),
            "so_luong": sl,
            "don_gia": _fmt_money(dg),
            "thanh_tien": _fmt_money(tt),
        })
    # Excel CTVT (cha + con, không giá)
    excel_rows = []
    for i, s in enumerate(services, start=1):
        excel_rows.append({
            "stt": i,
            "ten": (s.get("TEN_DV","") + SUFFIX_CHILD_LIST),
            "dvt": s.get("DVT",""),
            "so_luong": s.get("SO_LUONG", 0) or 0,
        })
        childs = [m for m in materials if (m.get("service_idx") in (i, str(i)) )]
        for j, m in enumerate(childs, start=1):
            excel_rows.append({
                "stt": f"{i}.{j}",
                "ten": m.get("TEN_VT",""),
                "dvt": m.get("DVT",""),
                "so_luong": m.get("SO_LUONG", 0) or 0,
            })
    # BBNT (cha có giá, con không giá)
    bbnt_rows = []
    for i, s in enumerate(services, start=1):
        sl = s.get("SO_LUONG", 0) or 0
        dg = s.get("DON_GIA", 0) or 0
        tt = sl * dg
        bbnt_rows.append({
            "stt": i,
            "ten": (s.get("TEN_DV","") + SUFFIX_CHILD_LIST),
            "dvt": s.get("DVT",""),
            "so_luong": sl,
            "don_gia": _fmt_money(dg),
            "thanh_tien": _fmt_money(tt),
        })
        childs = [m for m in materials if (m.get("service_idx") in (i, str(i)) )]
        for j, m in enumerate(childs, start=1):
            bbnt_rows.append({
                "stt": f"{i}.{j}",
                "ten": m.get("TEN_VT",""),
                "dvt": m.get("DVT",""),
                "so_luong": m.get("SO_LUONG", 0) or 0,
                "don_gia": "",
                "thanh_tien": "",
            })
    return contract_rows, excel_rows, bbnt_rows

# ---- DOCX ----
def render_docx(template_path: Path, ctx: Dict[str, Any], out_path: Path):
    tpl = DocxTemplate(template_path)
    tpl.render(ctx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tpl.save(out_path)

# ---- XLSX (dòng mẫu {{row.*}}) ----
_row_token_re = re.compile(r"\{\{\s*row\.[A-Za-z0-9_]+\s*\}\}")

from copy import copy
def _copy_row_styles(src_row, dst_row):
    for c_from, c_to in zip(src_row, dst_row):
        c_to.font = copy(getattr(c_from, "font", None))
        c_to.border = copy(getattr(c_from, "border", None))
        c_to.fill = copy(getattr(c_from, "fill", None))
        c_to.number_format = getattr(c_from, "number_format", "General")
        c_to.protection = copy(getattr(c_from, "protection", None))
        c_to.alignment = copy(getattr(c_from, "alignment", None))

from openpyxl import load_workbook
def render_excel_with_rows(template_path: Path, rows: List[Dict[str, Any]], out_path: Path):
    wb = load_workbook(template_path)
    ws = wb.active
    tpl_row_idx = None
    col_to_field = {}
    for r in ws.iter_rows():
        row_idx = r[0].row
        texts = [str(c.value) if c.value is not None else "" for c in r]
        if any(_row_token_re.search(t) for t in texts):
            tpl_row_idx = row_idx
            for c in r:
                if isinstance(c.value, str):
                    m = re.search(r"\{\{\s*row\.([A-Za-z0-9_]+)\s*\}\}", c.value or "")
                    if m:
                        col_to_field[c.column] = m.group(1)
            break
    if tpl_row_idx is None:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return

    tpl_cells = list(ws.iter_rows(min_row=tpl_row_idx, max_row=tpl_row_idx))[0]
    ws.delete_rows(tpl_row_idx, 1)

    cur = tpl_row_idx
    for rdata in rows:
        ws.insert_rows(cur, 1)
        new_cells = list(ws.iter_rows(min_row=cur, max_row=cur))[0]
        _copy_row_styles(tpl_cells, new_cells)
        for c in new_cells:
            field = col_to_field.get(c.column)
            if field:
                c.value = rdata.get(field, "")
        cur += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def render_excel_replace_tokens(template_path: Path, ctx: Dict[str, Any], out_path: Path):
    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    val = c.value
                    for k, v in ctx.items():
                        token = "{{"+k+"}}"
                        if token in val:
                            val = val.replace(token, str(v))
                    c.value = val
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

# ---- Bundle ----
def render_bundle(header: Dict[str,Any], services: List[Dict], materials: List[Dict], mode: str="draft"):
    if not header:
        raise RenderError("Thiếu header.")
    h = dict(header)

    for k in ["KH_TEN","KH_DIA_CHI","KH_DAI_DIEN","KH_CHUC_VU","DIA_DIEM"]:
        if h.get(k): h[k] = proper_case_vn(h[k])
    h = enrich_dates(h)

    if h.get("SO_HD_CORE"):
        year = dt.date.today().year
        if h.get("NGAY_HD_ISO"):
            year = dt.date.fromisoformat(h["NGAY_HD_ISO"]).year
        h["SO_HD"] = so_hd_full(h["SO_HD_CORE"], year)
        h["SO_HD_SEQ"] = h["SO_HD_CORE"]

    so_hd_disp = (h.get("SO_HD") or "").replace("/", ".")
    if not so_hd_disp:
        raise RenderError("Thiếu SO_HD (cần SO_HD_CORE + NGAY_HD).")
    h["SO_NGAY_HD"] = f"{h['SO_HD']}, {h.get('NGAY_HD_ISO','')}"

    if not h.get("BBNT_NGAY"):
        h["BBNT_NGAY"] = h.get("NGAY_KT") or h.get("NGAY_KT_FULL") or ""

    if h.get("KH_TEN"):
        h["KH_TEN_UPPER"] = h["KH_TEN"].upper()

    money = build_money(h, services)
    contract_rows, excel_rows, bbnt_rows = build_rows(services, materials)

    ctx = {}
    for k, v in h.items():
        ctx[k] = v
        ctx[k.lower()] = v
    ctx.update(money)
    ctx.update({
        "CONTRACT_ROWS": contract_rows,
        "EXCEL_ROWS": excel_rows,
        "BBNT_ROWS": bbnt_rows,
    })

    out_root = DRAFT_DIR / f"{so_hd_disp}_NHÁP" if mode=="draft" else FINAL_DIR
    out_root.mkdir(parents=True, exist_ok=True)

    mapping: List[Tuple[str, str, str]] = [
        ("1. HĐ",   "1. Mẫu hợp đồng cung cấp, lắp đặt_READY.docx", f"1.1 HĐ_{so_hd_disp}.docx"),
        ("2. CTVT", "2. Mẫu bảng chi tiết vật tư_READY.xlsx",      f"1.2 CTVT_{so_hd_disp}.xlsx"),
        ("3. BBNT", "3. Mẫu biên bản nghiệm thu_READY.docx",       f"1.3 BBNT&XNKL_{so_hd_disp}.docx"),
        ("4. BBTL", "4. Mẫu BB thanh lý HĐ_READY.docx",            f"1.4 BBTL_{so_hd_disp}.docx"),
        ("5. DNTT", "5. Mẫu đề nghị thanh toán_READY.docx",        f"1.5 DNTT_{so_hd_disp}.docx"),
    ]

    written = []
    missing = [str(TEMPLATES_DIR / src) for _, src, _ in mapping if not (TEMPLATES_DIR / src).exists()]
    if missing:
        raise RenderError("Thiếu template: " + " | ".join(missing))

    for _, src, dst in mapping:
        sp = TEMPLATES_DIR / src
        dp = out_root / dst
        if sp.suffix.lower() == ".docx":
            render_docx(sp, ctx, dp)
        else:
            render_excel_with_rows(sp, excel_rows, dp)
            render_excel_replace_tokens(dp, {k:str(v) for k,v in ctx.items() if not isinstance(v, (list, dict))}, dp)
        written.append(str(dp))

    return out_root, written
